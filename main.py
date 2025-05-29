import re
import streamlit as st
import pandas as pd
import os
from app import data_processing as dp
from app import ui
from app import config

st.set_page_config(page_title="Project Job board", layout="wide")
st.title("Project Job board")

# --- Carga de archivos ---
def get_file(path, file_uploader_label, type, key):
    uploaded = ui.file_uploader(file_uploader_label, type=type, key=key)
    if uploaded:
        return uploaded, True
    elif os.path.exists(path):
        return path, False
    else:
        st.warning(f"No se encontró el archivo por defecto en {path} y no se ha subido ninguno.")
        return None, False

candidates_file, candidates_uploaded = get_file(config.DEFAULT_CANDIDATES_PATH, "Sube archivo de candidatos (CSV)", ["csv"], "cand")
comments_file, comments_uploaded = get_file(config.DEFAULT_COMMENTS_PATH, "Sube archivo de comentarios (XLSX)", ["xlsx"], "comm")
template_file, template_uploaded = get_file(config.DEFAULT_TEMPLATE_PATH, "Sube template Teamtailor (XLSX)", ["xlsx"], "templ")

if not all([candidates_file, comments_file, template_file]):
    st.stop()

# --- Procesamiento de datos ---
import streamlit as st
if candidates_uploaded and str(candidates_file.name).lower().endswith('.xlsx'):
    candidates_df = pd.read_excel(candidates_file)
elif candidates_uploaded and str(candidates_file.name).lower().endswith('.csv'):
    candidates_df = pd.read_csv(candidates_file, sep=',', engine='python')
else:
    candidates_df = dp.read_candidates(candidates_file)


comments_df = pd.read_excel(comments_file) if comments_uploaded else dp.read_comments_excel(comments_file)
template_df = dp.read_template(template_file, sheet=config.SHEET_TEMPLATE)

# --- Merge de comentarios ---
comments_map = dp.merge_comments(comments_df)

# --- Construcción de tabla final ---
import json

def extract_linkedin(row):
    # Buscar primero columnas extraídas
    extracted_cols = [c for c in row.index if c.lower().replace(' ', '').startswith('linkedin') and c.endswith('_Extracted')]
    for col in extracted_cols:
        val = row.get(col)
        if pd.notnull(val):
            return str(val)
    # Si no, buscar en las columnas originales
    for col in ['14_Linkedin URL', 'Linkedin Url', 'LinkedIn', 'linkedin']:
        val = row.get(col)
        if pd.notnull(val):
            # Si es un hipervínculo de Excel (openpyxl lo puede leer como objeto), extraer string
            if hasattr(val, 'hyperlink') and val.hyperlink:
                return val.hyperlink.target
            # Si es un string JSON, intentar extraer el campo 'uri'
            val_str = str(val)
            if val_str.strip().startswith('{'):
                try:
                    js = json.loads(val_str)
                    if 'uri' in js:
                        return js['uri']
                except Exception:
                    pass
            return val_str
    return ''

def format_phone(phone):
    if pd.isna(phone) or not str(phone).strip():
        return ''
    # Remove all '+' characters from phone number
    return str(phone).replace('+', '').strip()

def format_tags(tags):
    if pd.isna(tags) or not str(tags).strip():
        return ''
    # Replace pipe (|) with semicolon (;)
    return '; '.join(tag.strip() for tag in str(tags).split('|') if tag.strip())

def is_valid_linkedin_url(url):
    """
    Validates if a URL is a properly formatted LinkedIn URL.
    Returns (is_valid, normalized_url)
    """
    if pd.isna(url) or not str(url).strip():
        return False, ''
    
    url = str(url).strip()
    
    # Check if it's a LinkedIn URL (case insensitive)
    if 'linkedin.com' not in url.lower():
        return False, ''
    
    # Handle common issues
    
    # 1. Handle multiple 'www' or mixed case 'Www' or 'WWW'
    url = re.sub(r'(?i)https?://(www\.)*', 'https://', url, 1)
    
    # 2. Ensure it starts with https://
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    
    # 3. Ensure it has www. after https://
    if url.startswith('https://') and not url.startswith('https://www.'):
        url = url.replace('https://', 'https://www.', 1)
    
    # 4. Force HTTPS
    if url.startswith('http://'):
        url = url.replace('http://', 'https://', 1)
    
    # 5. Remove any trailing slashes or parameters
    url = url.split('?')[0].split('#')[0].rstrip('/')
    
    # 6. Validate the domain pattern
    pattern = r'^https://www\.[a-z]{2,3}\.linkedin\.com/.*|^https://www\.linkedin\.com/.*'
    if not re.match(pattern, url.lower()):
        return False, ''
    
    # 7. Check for country codes in the URL
    country_codes = ['ar', 'au', 'at', 'bh', 'be', 'br', 'ca', 'cl', 'cn', 'co', 'cz', 'dk', 'fi', 'fr', 'de', 'gr',
                    'hk', 'hu', 'in', 'id', 'ie', 'il', 'it', 'jp', 'kr', 'kw', 'mx', 'my', 'nl', 'nz', 'no', 'pk',
                    'pe', 'ph', 'pl', 'pt', 'qa', 'ro', 'ru', 'sa', 'sg', 'za', 'es', 'se', 'ch', 'tw', 'th', 'tr',
                    'ae', 'uk', 'us', 'uy', 've', 'vn']
    
    # Check for country-specific URLs that don't match the pattern (e.g., uy.linkedin.com)
    country_pattern = r'^https?://([a-z]{2})\.linkedin\.com/.*'
    country_match = re.match(country_pattern, url.lower())
    if country_match and country_match.group(1) not in country_codes:
        return False, ''
    
    return True, url

def format_linkedin_url(url):
    """
    Format and validate a LinkedIn URL.
    Returns an empty string if the URL is invalid or not a LinkedIn URL.
    """
    is_valid, formatted_url = is_valid_linkedin_url(url)
    return formatted_url if is_valid else ''

def build_final_table(candidates_df, comments_map):
    rows = []
    for idx, row in candidates_df.iterrows():
        cid = row.get('id')
        # Split name into first and last for the template
        name = dp.clean_and_concat_names(row)
        first_name = ''
        last_name = ''
        if name:
            parts = name.strip().split()
            if len(parts) == 1:
                first_name = parts[0]
                last_name = ''
            elif len(parts) == 2:
                first_name = parts[0]
                last_name = parts[1]
            elif len(parts) == 3:
                first_name = ' '.join(parts[:2])
                last_name = parts[2]
            else:  # 4 or more
                first_name = ' '.join(parts[:2])
                last_name = ' '.join(parts[-2:])
        
        email = row.get('email')
        phone = format_phone(row.get('phone'))  # Format phone number
        download = row.get('download')  # Resume Name / CV Name
        tags = format_tags(row.get('tags'))  # Format tags
        linkedin = extract_linkedin(row)
        new_linkedin_url = format_linkedin_url(linkedin)  # New formatted LinkedIn URL
        
        note_raw = comments_map.get(cid, '')
        note = note_raw if note_raw else ''
        
        # Build row in the exact Teamtailor template order:
        rows.append({
            'Candidate Id': cid,
            'Email': email,
            'Resume Name / CV Name': download,
            'First Name': first_name,
            'Last Name': last_name,
            'Phone': phone,
            'Linkedin Url': linkedin,
            'new_linkedin_url': new_linkedin_url,  # New column
            'Tags': tags,
            'Note 1': note
        })
    
    # Ensure the DataFrame has the columns in the correct order
    columns_order = [
        'Candidate Id', 'Email', 'Resume Name / CV Name', 'First Name', 'Last Name',
        'Phone', 'Linkedin Url', 'new_linkedin_url', 'Tags', 'Note 1'
    ]
    
    return pd.DataFrame(rows)[columns_order]



final_df = build_final_table(candidates_df, comments_map)

import pandas as pd
import re

def format_notes(note):
    if pd.isna(note) or not note.strip():
        return ''
        
    # Split by double newlines and filter out empty comments
    comments = [c.strip() for c in note.split('\n\n') if c.strip()]
    
    # If no valid comments, return empty string
    if not comments:
        return ''
    
    # Process each comment to ensure it starts with 'note:'
    formatted_comments = []
    for comment in comments:
        # Remove any existing note prefix
        comment = re.sub(r'^note\s*:?\s*', '', comment, flags=re.IGNORECASE).strip()
        # Add the note prefix without numbering
        formatted_comments.append(f'note: {comment}')
    
    # Join all comments with double newlines
    return '\n\n'.join(formatted_comments)

# --- Selección de filas a exportar ---
st.subheader("Vista previa de la tabla final")
filtered_df = ui.row_selector(final_df)
# Apply note formatting for preview
filtered_df = filtered_df.copy()
filtered_df['Note 1'] = filtered_df['Note 1'].apply(format_notes)
ui.show_table(filtered_df)

# --- Botón para descargar template ---
def to_excel(df):
    import io
    from openpyxl import load_workbook
    import pandas as pd
    import numpy as np
    # Clean DataFrame: fill NaN/None, convert all to string, and remove invalid XML characters
    import re
    
    def clean_xml_chars(val):
        if not isinstance(val, str):
            val = str(val)
        # Remove all control characters except tab, CR, LF, and only allow valid XML unicode ranges
        return re.sub(r'[^\x09\x0A\x0D\x20-\uD7FF\uE000-\uFFFD]', '', val)
    
    # Clean the DataFrame
    df_clean = df.fillna('').replace({None: ''})
    df_clean = df_clean.astype(str)
    df_clean = df_clean.applymap(clean_xml_chars)

    # Load template and get headers
    wb = load_workbook(template_file)
    ws = wb.active
    template_headers = [str(cell.value).strip() for cell in ws[1] if cell.value is not None]
    
    # Get the intersection of template headers and DataFrame columns
    # This ensures we only include columns that exist in both
    common_headers = [h for h in template_headers if h in df_clean.columns]
    
    # If we have new columns in the DataFrame that aren't in the template,
    # we'll add them to the end
    new_columns = [col for col in df_clean.columns 
                  if col not in template_headers]
    
    # If there are new columns, we need to add them to the template
    if new_columns:
        # Find the last column with data in the first row
        last_col = ws.max_column
        for col_idx, col_name in enumerate(new_columns, start=last_col + 1):
            # Add the new column header
            ws.cell(row=1, column=col_idx, value=col_name)
            # Update template headers
            template_headers.append(col_name)
    
    # Reorder DataFrame columns to match template + any new columns
    output_headers = [h for h in template_headers if h in df_clean.columns]
    df_export = df_clean[output_headers].copy()
    
    # Update the template headers in the worksheet if we added new columns
    if new_columns:
        for idx, header in enumerate(template_headers, 1):
            ws.cell(row=1, column=idx, value=header)

    # Remove previous data (keep headers)
    ws.delete_rows(2, ws.max_row)  # Delete all rows except the header

    # Append data row by row
    for _, row in df_export.iterrows():
        ws.append([row.get(h, '') for h in template_headers])
    
    # Save to in-memory bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Streamlit download button for normal export

# Apply the same formatting for export
export_df = filtered_df.copy()
export_df['Note 1'] = export_df['Note 1'].apply(format_notes)
excel_bytes = to_excel(export_df)
if excel_bytes:
    st.download_button(
        label="Descargar Excel",
        data=excel_bytes,
        file_name="Teamtailor_Export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --- Table: Candidates with more than one comment ---
# First, count comments before formatting
main_df = filtered_df.copy()

def count_comments(note):
    if pd.isna(note) or not note.strip():
        return 0
    return len([c for c in note.split('\n\n') if c.strip()])

# Add comment count before formatting
main_df['comment_count'] = main_df['Note 1'].apply(count_comments)

# Format notes after counting
main_df['Note 1'] = main_df['Note 1'].apply(format_notes)

# Show only candidates with more than one comment
multi_comment_df = main_df[main_df['comment_count'] > 1]
st.subheader("Candidatos con más de un comentario")
st.dataframe(multi_comment_df[['Candidate Id', 'First Name', 'Last Name', 'Email', 'comment_count', 'Note 1']])
#     file_name="Teamtailor_Import_Template.xlsx",
#     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# )
