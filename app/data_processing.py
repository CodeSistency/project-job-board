"""
Funciones para el procesamiento, validación y merge de datos de candidatos y comentarios.
"""
import pandas as pd
from typing import Optional, List, Dict
import os

def extract_linkedin_urls_xlsx(path: str, linkedin_columns=None) -> pd.DataFrame:
    """
    Lee un archivo Excel y extrae los hyperlinks de las columnas de LinkedIn si existen.
    Devuelve un DataFrame con una columna adicional 'LinkedIn_Extracted' si se encuentra.
    """
    import openpyxl
    if linkedin_columns is None:
        linkedin_columns = ['14_Linkedin URL', 'Linkedin Url', 'LinkedIn', 'linkedin']

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    data = list(ws.values)
    headers = [str(h) if h is not None else "" for h in data[0]]
    df = pd.DataFrame(data[1:], columns=headers)

    # Extract hyperlinks
    for col in linkedin_columns:
        if col in ws[1]:
            col_idx = headers.index(col) + 1  # openpyxl is 1-based
            urls = []
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                if cell.hyperlink:
                    urls.append(cell.hyperlink.target)
                else:
                    urls.append(cell.value)
            df[col + '_Extracted'] = urls
    return df

def read_candidates(path: str) -> pd.DataFrame:
    """Lee el archivo de candidatos (CSV) y retorna un DataFrame. Extrae LinkedIn URL del campo JSON si está presente."""
    import pandas as pd
    df = pd.read_csv(path, sep=',', engine='python')
    return df

def read_comments_excel(path: str) -> pd.DataFrame:
    """Lee el archivo de comentarios Excel y retorna un DataFrame."""
    return pd.read_excel(path)

def read_template(path: str, sheet: str = 'Candidates') -> pd.DataFrame:
    """Lee la hoja 'Candidates' del template y retorna un DataFrame."""
    return pd.read_excel(path, sheet_name=sheet)

def merge_comments(comments_df: pd.DataFrame) -> Dict:
    """
    Agrupa y concatena comentarios por candidate_id en el formato requerido.
    Retorna un dict: {candidate_id: comentarios_concatenados}
    """
    result = {}
    grouped = comments_df.groupby('candidate_id')
    for cid, group in grouped:
        comentarios = []
        for _, row in group.iterrows():
            comentarios.append(f"{row['id']} {row['created_at']}: {row['body']}")
        result[cid] = '\n\n'.join(comentarios)
    return result

def clean_and_concat_names(row: pd.Series) -> str:
    """
    Concatena y limpia nombres y apellidos según reglas del PRD.
    """
    parts = []
    # Ajustar según los nombres de columnas en el CSV
    for col in ['name', 'lastname', 'basic', 'First Name', 'Last Name']:
        if col in row and pd.notnull(row[col]):
            parts.append(str(row[col]).strip())
    return ' '.join(parts).replace('  ', ' ')

# Más funciones a agregar según avance del desarrollo
